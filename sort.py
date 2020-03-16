import numpy as np
import pandas as pd
import os
import matplotlib.pyplot as plt
from copy import deepcopy as copy
import inputData as inputData
from numpy import random
from collections import Set
import sys
from openpyxl import load_workbook
import math

sys.setrecursionlimit(80000)
variables = []
var_dict = dict()
size_of_group = 5
num_plot = 131
num_iterations = 1
max_samples = int(3e3)  # How many iterations to run the algorithm for (500-5000 works best)
costThreshold = 1950    # Experiement with different threshold values depending on weights used
name_of_excel = "exampleList"
current_directory = os.getcwd()
excelOut = os.path.join(current_directory, 'Sorted_Output.xlsx') # Path of the Excel sheet that we're outputting data to

#these can be modified by hand or a program could be written to modify and test for best outcome
weights = {
    "gender" : 7,
    "residency" : 5,
    "ethnicity" : 6,
    "major" : 3,
    "rural" : 3,
    "public" : 2
    }

class Person:
    def __init__(self, name):
        self.name = name
        self.gender = ""
        self.residency = ""
        self.region = ""
        self.school = ""
        self.ethnicity = ""
        self.major = ""
        self.availability = ""
        self.incompatable = ""
        self.public = ""
        self.rural = ""

        self.dict = {
            "name" : self.name,
            "gender" : self.gender,
            "residency" : self.residency,
            "school" : self.school,
            "ethnicity" : self.ethnicity,
            "major" : self.major,
            "availability" : self.availability,
            "incompatable" : self.incompatable,
            "public" : self.public,
            "rural" : self.rural,
            "region" : self.region
        }

    #Turns strings of data into lists and stores in respective field
    def cleanData(self):
        for var in variables:
                try:
                    if "," in self.dict[var]:
                        self.dict[var] = self.dict[var].split(", ")
                except:
                    continue



    def printPerson(self):
        print(f"Name: {self.name}", end = "  ")
        for var in variables:
            print(f"{var}: {self.dict[var]}", end = "  ")
        print("")

class Group:
    def __init__(self, people = []):
        self.people = people

        self.funcDict = {
            "gender" : self.binaryCost,
            "rural" : self.binaryCost,
            "public" : self.binaryCost,
            "residency" : self.diverseCost,
            "region" : self.binaryCost,
            "ethnicity" : self.diverseCost,
            "major" : self.majorCost,
        }

    #Prints the cost, ocombining each individual variable's cost
    def groupCost(self):
        cost = 0
        for var in variables:
            try:
                cost += self.funcDict[var](var) * weights[var]
            except:
                continue
        return cost

    #We want the largest number of majors present
    def majorCost(self, major):
        dead_science = 0
        life_science = 0
        humanities = 0
        finance = 0
        social_science = 0
        cost = 5
        for person in self.people:
            for i in range(1):
                major = person.dict["major"]
                if("Computer Science" or "Mathematics" or "Physics" or "Statistics" or "Chemistry" in major):
                    dead_science += 1
                elif("Biology" or "Neuroscience" or "Nutrition" or "Medical Anthropology" or "Environmental Health Sciences" or "Biomedical Engineering" in major):
                    life_science += 1
                elif("Undecided" or "Philosophy" or "Journalism" or "Communication" or "Spanish Literature and Cultures" or "History" in major):
                    humanities += 1
                elif("Business" or "Economics" in major):
                    finance += 1
                elif("PWAD" or "Public Policy" or "Political Science" or "Pscyhology" or "Cognitive Science" or "Health Policy and Management" or "Human Development and Family Studies" or "Global Studies" in major):
                    social_science += 1
                else:
                    print("YOU FORGOT A MAJOR!")
        if(dead_science > 0):
            cost -- 1
        if(life_science > 0):
            cost -- 1
        if(humanities > 0):
            cost -- 1
        if(finance > 0):
            cost -- 1
        if(social_science > 0):
            cost -- 1
        return cost

    def diverseCost(self, variable):
        diverse = set()
        for person in self.people:
            try:
                diverse.add(person.dict[variable])
            except:
                for i in person.dict[variable]:
                    diverse.add(i)
        if len(diverse) == 1:  # Heavily weight against having all the same ethnicity (which happens a lot)
            return 10

        return size_of_group - len(diverse)

    def binaryCost(self, attribute):
        attr_one = ""
        num_attr_one = 0
        num_attr_two = 0

        for person in self.people:
            if attr_one == "":
                attr_one = person.dict[attribute]
                num_attr_one += 1
            elif person.dict[attribute] == attr_one:
                num_attr_one += 1
            else:
                num_attr_two += 1

        else:   # Otherwise optimize for simple heterogeneity
            # Since groups are of five people, accept an imbalance if there's a discrepancy of 1
            if abs(num_attr_one - num_attr_two) <= 1:
                return 0
            else:
                return abs(num_attr_one - num_attr_two)

    #Returns true if there are no two scholars from the same school
    def isValidGroup(self):
        # Store all the schools and regions present in the group
        for var in variables:
            if var == "incompatable":
                if not self.checkCompatability():
                    #print("not compatable")
                    return False
            if var == "school":
                if not self.checkSchool():
                    #print("same school")
                    return False
            if var == "availability":
                if not self.checkAvailability():
                    #print("no availability")
                    return False

        return True

    def checkCompatability(self):
        for student in self.people:
            for incompatable in student.incompatable:
                for checkStudent in self.people:
                    if(incompatable == checkStudent):
                        return False
        return True

    def checkSchool(self):
        schoolSet = set([])
        for i in range(len(self.people)):
            schoolSet.add(self.people[i].school)    # Always add the school

        # If there are any overlaps, then the length of each set will have decreased
        if len(schoolSet) < len(self.people): # or len(regionSet) < in_states:
            return False
        return True

    def checkAvailability(self):
        num_availability = len(self.people[0].dict["availability"])
        combined_availability = [0 for i in range(size_of_group)]

        for person in self.people:
            for i in range(4):
                if person.dict["availability"][i] == "T":
                    combined_availability[i] += 1
        cost = 4
        for i in range(size_of_group):
            if combined_availability[i] >= size_of_group:
                cost -- 1
        return cost

    def countEverything(self):
        males = 0
        females = 0
        in_states = 0
        out_of_states = 0
        rurals = 0
        not_rurals = 0
        publics = 0
        privates = 0
        ethnicities = set([])
        regions = set([])
        for person in self.people:
            ethnicities.add(person.ethnicity)
            regions.add(person.region)
            if person.gender == "True":
                males += 1
            else:
                females += 1
            if person.in_state == "True":
                in_states += 1
            else:
                out_of_states += 1
            if person.rural == "True":
                rurals += 1
            else:
                not_rurals += 1
            if person.public == "True":
                publics += 1
            else:
                privates += 1

        print(str(males) + " males, " + str(females) + " females")
        print(str(in_states) +  " in-states, " + str(out_of_states) + " out-of-states")
        print(str(rurals) + " rurals, " + str(not_rurals) + " not-rurals")
        print(str(publics)  + " publics, " + str(privates) + " privates ")
        print(str(len(ethnicities)) + " different ethnicities")
        print(str(len(regions)) + " different regions")

    def printGroup(self):
        print("GROUP COSTS:")
        for person in self.people:
            person.printPerson()
        for var in variables:
            try:
                print(f"{var} cost: {self.funcDict[var]}")
            except:
                continue
        print(f"Total Cost: {self.groupCost()}")
        print("")

def makeStartingGroups():

    data = pd.read_excel(dataPath)
    people = []
    groups = []

    #This loop creates all the people from the excel
    for index, row in data.iterrows():
        person = Person(row[var_dict["name"]])
        for var in variables:
            person.dict[var] = row[var_dict[var]]
        person.cleanData()
        people.append(person)
    all_indeces = np.arange(0, len(people), 1)
    print(len(all_indeces))

    #This loop creates groups one at a time
    currentGroupMembers = []
    currentIdxs = []
    for counter in range(len(all_indeces)):
        idx = random.choice(all_indeces)        # Select a random person index
        all_indeces = all_indeces[all_indeces != idx]
        people[idx].printPerson()
        currentGroupMembers.append(people[idx])
        currentIdxs.append(idx)
        if len(currentGroupMembers) == size_of_group:           # Once the group is full, register it and repeat
            newGroup = Group(currentGroupMembers)
            if newGroup.isValidGroup():
                print("Valid Group Made!")
                groups.append(newGroup)
                #all_indeces = all_indeces[all_indeces != any(currentIdxs)]             # Remove the current selection from the possibilities
                currentIdxs = []
                currentGroupMembers = []
            else:
                print("inValidGroup")
                counter -= size_of_group
                currentGroupMembers = []
                currentIdxs = []
                all_indeces = np.concatenate(all_indeces, currentIdxs)


    return groups

    print("No valid set of groups could be made")


# In[1]:


# Gets the total cost for a list of groups
def totalCost(groups):
    sum = 0
    for group in groups:
        sum += group.groupCost()
    return sum


# Use a simmulated annealing algorithm to optimize the group placement
def sort(groups, max_samples):
    print('Running Simmulated Annealing Algorithm')
    cost_history = np.zeros(max_samples)    # Store the total cost at each time step to see if this actually works
    sample_num = 0
    min_cost = np.inf   # Anything will be better than this starting cost
    best_group = None
    best_iteration = 0

    # Variable that decreases over the duration that corresponds to jump probability
    temp = 5                  # Experiment with different starting temperature values
    dT = temp / max_samples    # Amount by which the temperature decreases at each step (linearly)

    while sample_num < max_samples:
        currentCost = totalCost(groups)
        cost_history[sample_num] = currentCost
        if currentCost < min_cost:  # See if this is the best group ever and save if so
            bestGroup = copy(groups)
            min_cost = currentCost
            best_iteration = sample_num

        testGroups = makeSwap(groups)
        nextCost = totalCost(testGroups)

        if nextCost <= currentCost:  # If the cost decreases, it's guaranteed to be a good move!
            groups = testGroups

        else:     # If the cost increases, accept the change with a random probability
            u = random.uniform(0, 1)                                # Get a random variable from a uniform distribution
        #     # We only end up here if nextCost > currentCost, so the exponentiated term is always negative
        #     # High acceptance when temperature is HIGH or when the next cost is only slightly worse
            acceptance = np.exp((currentCost - nextCost) / temp)    # High if the cost decreases or temperature is low
            if acceptance >= u:  # Accept if it's above the random threshold
                groups = testGroups
        sample_num += 1
        temp -= dT                          # Decrease the temperature

    print('The Best Group Was Found At Iteration {} with Cost {}'.format(best_iteration, min_cost))
    return bestGroup, cost_history

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, header=True,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass
    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, header=header, startrow=startrow, **to_excel_kwargs, index=False)

    # save the workbook
    writer.save()


# In[ ]:


def exportGroups(final_groups, excelOut):
    print("EXCEL OUTTED")
    # Once we decide on a good final group, export it to a spreadsheet
    for i, group in enumerate(final_groups, 1):
        sheet_name = 'Group {}'.format(i)
        rows = []

        for person in group.people:
            costRow = ["Cost"]
            properties = [person.dict["name"]]
            for var in variables:
                properties.append(person.dict[var])
                try:
                    costRow.append(group.funcDict[var](var))
                except:
                    continue

            rows.append(properties)

        originalColumns = copy(columns)
        originalRows = copy(rows)   # Rows without the group number attached (for use in the group-specific sheet)
        for row in rows:
            row.append(i)
        newColumns = np.append(columns, "Group #")
        totalDF = pd.DataFrame(data=rows, columns=newColumns)

        append_df_to_excel(excelOut, totalDF, header=(i==1), sheet_name='All Scholars')

        #costRow.append(group.groupCost())
        originalRows.append(costRow)
        groupDF = pd.DataFrame(data=originalRows, columns=originalColumns)
        append_df_to_excel(excelOut, groupDF, sheet_name='Group {}'.format(i))

# Swaps two random people in random groups
def makeSwap(input_groups):
    #print("CALLED MAKE SWAP")
    #print(f"Beginning cost: {totalCost(input_groups)}")
    groups = copy(input_groups) #
    rankedGroups = sorted(groups, key=lambda group: group.groupCost(), reverse=True)
    worstGroup = rankedGroups[int(np.random.normal(loc=1, scale=1))]
    # Pick the second group randomly
    group2 = rankedGroups[random.randint(1, len(groups))]

    # Use copies to avoid accidentally overwriting
    def randomSwap(worstGroup, group2):
        person1_idx = np.random.randint(0, len(worstGroup.people))
        person2_idx = np.random.randint(0, len(group2.people))
        person1 = copy(worstGroup.people[person1_idx])
        person2 = copy(group2.people[person2_idx])

        # Make the swap
        worstGroup.people[person1_idx] = person2
        group2.people[person2_idx] = person1

        # Only allow the swap if both resulting groups are still valid (no overlapping schools)
        if (not worstGroup.isValidGroup() or not group2.isValidGroup()):
            worstGroup.people[person1_idx] = person1    # Undo the change and repeat
            group2.people[person2_idx] = person2
            randomSwap(worstGroup, group2)              # If not valid, just repeat and hope for different random indeces

    randomSwap(worstGroup, group2)

    return rankedGroups

def plotGroup(input_groups):
    groups = copy(input_groups)
    global num_plot
    names = []
    for i in range(len(groups)):
        names.append(f"Group {i}")
    values = [group.groupCost() for group in groups]

    average = np.average(values)

    plt.subplot(num_plot)
    num_plot += 1
    plt.bar(names, values)
    plt.xticks([])
    plt.axhline(y=average,linewidth=1, color='k')

def run():
    global num_plot
    plt.figure()
    # Make sure we're using a new Excel output
    try:
        os.remove(excelOut)
    except FileNotFoundError:
        pass

    max_samples = int(3e3)  # How many iterations to run the algorithm for (500-5000 works best)
    costThreshold = 1950    # Experiement with different threshold values depending on weights used


    start_groups = makeStartingGroups()     # Start with a totally randomized grouping
    plotGroup(start_groups)
    initialCost = int(totalCost(start_groups))
    print(f"Final cost: {initialCost}")

    plt.subplot(num_plot)
    num_plot += 1

    final_groups, final_history = sort(start_groups, 1000)
    for i in range(num_iterations):
        temp_groups, temp_history = sort(start_groups, 1000)
        plt.plot(temp_history)
        if totalCost(temp_groups) < totalCost(final_groups):
            final_groups = temp_groups
            #final_history = temp_history
    plt.xlabel('Sample Number')
    plt.ylabel('Total Cost Function')
    plt.title('Simmulated Annealing Cost Optimization')
    plotGroup(final_groups)
    finalCost = int(totalCost(final_groups))
    print(f"Final cost: {finalCost}")
    exportGroups(final_groups, excelOut)
    return final_groups







# Make sure we're using a new Excel output
try:
    os.remove(excelOut)
except FileNotFoundError:
    pass

name_of_excel, size_of_group, max_samples, num_iterations = inputData.askForInput(name_of_excel, size_of_group, max_samples, num_iterations)
print(name_of_excel, size_of_group, max_samples, num_iterations)
dataPath = os.path.join(current_directory, f"{name_of_excel}.xlsx") #default path

data = pd.read_excel(dataPath)

columns = data.columns.ravel()

print(f"{variables}")
for var in columns:
    print(var)
    if any(word in var for word in ["number", "name", "Name", "Number"]):
        var_dict["name"] = var
    if any(word in var for word in ["gender", "Gender", "sex", "Sex"]):
        var_dict["gender"] = var
        variables.append("gender")
    if any(word in var for word in ["ethnicity", "Ethnicity", "race", "Race"]):
        variables.append("ethnicity")
        var_dict["ethnicity"] = var
    if any(word in var for word in ["residency", "Residency"]):
        variables.append("residency")
        var_dict["residency"] = var
    if any(word in var for word in ["region", "Region"]):
        variables.append("region")
        var_dict["region"] = var
    if any(word in var for word in ["major", "Major"]):
        variables.append("major")
        var_dict["major"] = var
    if any(word in var for word in ["rural", "Rural", "urban", "Urban"]):
        variables.append("rural")
        var_dict["rural"] = var
    if any(word in var for word in ["public", "Public"]):
        variables.append("public")
        var_dict["public"] = var
    if any(word in var for word in ["Roommate", "roommate", "incompatable", "Incompatable"]):
        variables.append("incompatable")
        var_dict["incompatable"] = var
    if any(word in var for word in ["school", "School", "high school", "High School"]):
        variables.append("school")
        var_dict["school"] = var
    if any(word in var for word in ["Available", "available", "availability", "Availability"]):
        variables.append("availability")
        var_dict["availability"] = var

run()

plt.show()
